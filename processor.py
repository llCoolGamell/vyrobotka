"""
Логика обработки сводного запроса.

Вход: Excel-файл "сводный запрос" с колонками:
    Исполнитель | Роль | Киз | Boxcount_короба | Processcount_строки | Строк
(также принимается формат "выработка": Киз | короба | сборка_строк | контроль_строк)

Выход: на каждого сотрудника отдельный .xlsx с колонками формата "выработка":
    Исполнитель | Роль | Киз | короба | сборка_строк | контроль_строк | объём выполненных работ

Логика:
  - строки с одинаковыми (Исполнитель, Роль) суммируются;
  - числовые колонки складываются;
  - роли из EXCLUDED_ROLES полностью исключаются;
  - для строк из DUPLICATE_RULES внизу добавляется дубль с новым названием
    и тем же значением Киз (входит в ИТОГО);
  - "объём выполненных работ" = сумма всех числовых колонок строки;
  - сортировка по ФИО, затем по роли;
  - в конце файла каждого сотрудника добавляется итоговая строка "ИТОГО".
"""

import os
import re
import pandas as pd

# Для каждой итоговой колонки — возможные названия во входном файле
COLUMN_ALIASES = {
    "Киз": ["Киз"],
    "короба": ["Boxcount_короба", "короба"],
    "сборка_строк": ["Processcount_строки", "сборка_строк"],
    "контроль_строк": ["Строк", "контроль_строк"],
}
NUMERIC_OUT = ["Киз", "короба", "сборка_строк", "контроль_строк"]
OUT_COLUMNS = ["Исполнитель", "Роль"] + NUMERIC_OUT + ["объём выполненных работ"]

# Виды работ, которые полностью исключаются (нет в файлах и не в итогах)
EXCLUDED_ROLES = {
    "ТСД Льгота с упак.МАРК.товар,сборка штук",
    "ТСД ХОЛОД 2 сборка штук",
}

# Правила дублирования: исходная работа -> название новой строки-дубля.
# В дубль копируется значение колонки Киз из исходной строки (остальное 0).
# Дубли добавляются внизу (перед ИТОГО) и входят в общий итог.
DUPLICATE_RULES = [
    ("Сборка и упаковка в Холодильнике", "упаковка холод"),
    ("Сборка и упаковка на льготе", "упаковка марк"),
]
DUPLICATE_SOURCE_COLUMN = "Киз"


def _norm(s: str) -> str:
    """Нормализовать строку для сравнения: схлопнуть пробелы."""
    return " ".join(str(s).split())


_EXCLUDED_NORM = {_norm(r) for r in EXCLUDED_ROLES}
_DUPLICATE_NORM = [(_norm(src), dst) for src, dst in DUPLICATE_RULES]


def _safe_filename(name: str) -> str:
    """Убрать недопустимые для имени файла символы."""
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    return name[:150] or "без_имени"


def load_summary(path: str) -> pd.DataFrame:
    """Прочитать входной файл и привести колонки к выходному формату."""
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]

    missing = [c for c in ["Исполнитель", "Роль"] if c not in df.columns]

    rename_map = {}
    for out_name, aliases in COLUMN_ALIASES.items():
        found = next((a for a in aliases if a in df.columns), None)
        if found is None:
            missing.append(out_name + " (" + " / ".join(aliases) + ")")
        else:
            rename_map[found] = out_name

    if missing:
        raise ValueError(
            "В файле не найдены колонки: " + ", ".join(missing) +
            "\nНайдены: " + ", ".join(df.columns)
        )

    df = df.rename(columns=rename_map)

    # Убрать строки без ФИО или без роли (итоговые/пустые строки исходника)
    df = df[df["Исполнитель"].notna() & df["Роль"].notna()].copy()
    df["Исполнитель"] = df["Исполнитель"].astype(str).str.strip()
    df["Роль"] = df["Роль"].astype(str).str.strip()
    df = df[(df["Исполнитель"] != "") & (df["Роль"] != "")]

    # Исключить запрещённые виды работ полностью
    df = df[~df["Роль"].map(_norm).isin(_EXCLUDED_NORM)]

    # Числа -> числовой тип
    for c in NUMERIC_OUT:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    return df


def aggregate(df: pd.DataFrame) -> pd.DataFrame:
    """Сгруппировать одинаковые (ФИО, Роль) и просуммировать."""
    grouped = (
        df.groupby(["Исполнитель", "Роль"], as_index=False)[NUMERIC_OUT]
        .sum()
    )
    grouped["объём выполненных работ"] = grouped[NUMERIC_OUT].sum(axis=1)
    grouped = grouped.sort_values(["Исполнитель", "Роль"]).reset_index(drop=True)
    return grouped[OUT_COLUMNS]


def _add_duplicates(emp_df: pd.DataFrame, employee: str) -> pd.DataFrame:
    """Добавить внизу строки-дубли по DUPLICATE_RULES (значение из Киз)."""
    roles_norm = emp_df["Роль"].map(_norm)
    extra = []
    for src_norm, dst_name in _DUPLICATE_NORM:
        match = emp_df[roles_norm == src_norm]
        if match.empty:
            continue
        value = match[DUPLICATE_SOURCE_COLUMN].sum()
        row = {c: 0 for c in NUMERIC_OUT}
        row[DUPLICATE_SOURCE_COLUMN] = value
        row["Исполнитель"] = employee
        row["Роль"] = dst_name
        row["объём выполненных работ"] = value
        extra.append(row)
    if not extra:
        return emp_df
    return pd.concat([emp_df, pd.DataFrame(extra)[OUT_COLUMNS]],
                     ignore_index=True)


def _write_employee_file(emp_df: pd.DataFrame, out_path: str):
    """Записать .xlsx одного сотрудника с итоговой строкой и форматированием."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    emp_df = emp_df.copy()

    totals = {c: emp_df[c].sum() for c in NUMERIC_OUT + ["объём выполненных работ"]}
    totals["Исполнитель"] = "ИТОГО"
    totals["Роль"] = ""
    emp_df = pd.concat([emp_df, pd.DataFrame([totals])[OUT_COLUMNS]],
                       ignore_index=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        emp_df.to_excel(writer, index=False, sheet_name="Выработка")
        ws = writer.sheets["Выработка"]

        header_fill = PatternFill("solid", fgColor="2E7D32")
        header_font = Font(color="FFFFFF", bold=True)
        total_fill = PatternFill("solid", fgColor="C8E6C9")
        thin = Side(style="thin", color="BBBBBB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        last_row = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=last_row):
            for cell in row:
                cell.border = border
        for cell in ws[last_row]:
            cell.fill = total_fill
            cell.font = Font(bold=True)

        widths = [32, 42, 10, 10, 14, 16, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

    return out_path


def process(input_path: str, output_dir: str, progress=None) -> list:
    """
    Полный цикл: чтение -> агрегирование -> файлы по сотрудникам.
    progress(done, total, name) — необязательный колбэк.
    Возвращает список путей созданных файлов.
    """
    os.makedirs(output_dir, exist_ok=True)
    df = load_summary(input_path)
    agg = aggregate(df)

    employees = list(agg["Исполнитель"].unique())
    total = len(employees)
    created = []

    for i, emp in enumerate(employees, start=1):
        emp_df = agg[agg["Исполнитель"] == emp].reset_index(drop=True)
        emp_df = _add_duplicates(emp_df, emp)
        fname = _safe_filename(emp) + ".xlsx"
        out_path = os.path.join(output_dir, fname)
        _write_employee_file(emp_df, out_path)
        created.append(out_path)
        if progress:
            progress(i, total, emp)

    return created


if __name__ == "__main__":
    import sys
    inp = sys.argv[1]
    outd = sys.argv[2] if len(sys.argv) > 2 else "out"
    files = process(inp, outd, progress=lambda d, t, n: print(f"[{d}/{t}] {n}"))
    print(f"Готово. Создано файлов: {len(files)}")
